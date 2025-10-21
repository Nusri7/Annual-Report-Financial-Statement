import os
import time
import requests
import pandas as pd
from io import StringIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from dotenv import load_dotenv
import re
from unstract.llmwhisperer import LLMWhispererClientV2
from unstract.llmwhisperer.client_v2 import LLMWhispererClientException
# sop_engine usage removed from main flow

# Load environment variables from .env file
load_dotenv()

# =====================
# CONFIG
# =====================
# Load API keys from .env file
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

# Load all three LLMWhisperer API keys
LLMWHISPERER_API_KEYS = [
    os.getenv("LLMWHISPERER_API_KEY_1"),
    os.getenv("LLMWHISPERER_API_KEY_2"), 
    os.getenv("LLMWHISPERER_API_KEY_3")
]

# Filter out None/empty keys and validate
LLMWHISPERER_API_KEYS = [key for key in LLMWHISPERER_API_KEYS if key and key.strip()]

# Validate required environment variables
if not OPENROUTER_API_KEY:
    raise ValueError("OPENROUTER_API_KEY not found in .env file")
if not LLMWHISPERER_API_KEYS:
    raise ValueError("No valid LLMWHISPERER_API_KEY found in .env file. Please set at least LLMWHISPERER_API_KEY_1")

print(f"🔑 Loaded {len(LLMWHISPERER_API_KEYS)} LLMWhisperer API key(s)")

MODEL = "x-ai/grok-code-fast-1"
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"
HEADERS = {
    "Authorization": f"Bearer {OPENROUTER_API_KEY}",
    "Content-Type": "application/json"
}

# =====================
# LLMWhisperer API Key Manager
# =====================
class LLMWhispererKeyManager:
    """Manages rotation of LLMWhisperer API keys when daily limits are hit"""
    
    def __init__(self, api_keys):
        self.api_keys = api_keys
        self.current_key_index = 0
        self.failed_keys = set()  # Track which keys have failed today
        self.successful_keys = set()  # Track which keys have worked
        self.total_attempts = 0
        
    def get_current_key(self):
        """Get the current API key to use"""
        if self.current_key_index < len(self.api_keys):
            return self.api_keys[self.current_key_index]
        return None
    
    def mark_key_failed(self, reason="Daily limit exceeded"):
        """Mark current key as failed and rotate to next"""
        current_key = self.get_current_key()
        if current_key:
            self.failed_keys.add(current_key)
            print(f"❌ API key {self.current_key_index + 1} failed: {reason}")
        
        self.current_key_index += 1
        if self.current_key_index < len(self.api_keys):
            print(f"🔄 Rotating to API key {self.current_key_index + 1}")
            print(f"📊 Key Status: {self.get_key_status()}")
            return True
        else:
            print(f"❌ All {len(self.api_keys)} API keys have been exhausted")
            print(f"📊 Final Status: {self.get_key_status()}")
            return False
    
    def mark_key_successful(self):
        """Mark current key as successful"""
        current_key = self.get_current_key()
        if current_key:
            self.successful_keys.add(current_key)
            print(f"✅ API key {self.current_key_index + 1} is working successfully")
    
    def has_available_keys(self):
        """Check if there are still available keys to try"""
        return self.current_key_index < len(self.api_keys)
    
    def get_key_status(self):
        """Get detailed status of all keys for logging"""
        status = []
        for i, key in enumerate(self.api_keys):
            if i < self.current_key_index:
                status.append(f"Key {i+1}: ❌ Failed")
            elif i == self.current_key_index:
                status.append(f"Key {i+1}: 🔄 Current")
            else:
                status.append(f"Key {i+1}: ⏳ Available")
        return " | ".join(status)
    
    def get_summary(self):
        """Get a summary of key usage"""
        return {
            'total_keys': len(self.api_keys),
            'current_key': self.current_key_index + 1 if self.current_key_index < len(self.api_keys) else None,
            'failed_keys': len(self.failed_keys),
            'successful_keys': len(self.successful_keys),
            'remaining_keys': max(0, len(self.api_keys) - self.current_key_index),
            'total_attempts': self.total_attempts
        }

# Initialize the key manager
key_manager = LLMWhispererKeyManager(LLMWHISPERER_API_KEYS)

# =====================
# Helper: Call OpenRouter API
# =====================
def call_openrouter(messages):
    print(f"🤖 Calling OpenRouter API with model: {MODEL}")
    payload = {"model": MODEL, "messages": messages, "temperature": 0}
    
    try:
        r = requests.post(OPENROUTER_URL, headers=HEADERS, json=payload, timeout=120)
        r.raise_for_status()
        response = r.json()
        content = response["choices"][0]["message"]["content"]
        print(f"✅ API call successful. Response length: {len(content)} characters")
        return content
    except requests.exceptions.RequestException as e:
        print(f"❌ API request error: {e}")
        return ""
    except KeyError as e:
        print(f"❌ Unexpected API response format: {e}")
        return ""
    except Exception as e:
        print(f"❌ Unexpected error in API call: {e}")
        return ""

def call_openrouter_custom_model(messages, model: str, temperature: float = 0):
    """Call OpenRouter with a user-specified model and temperature."""
    print(f"🤖 Calling OpenRouter API with custom model: {model}")
    payload = {"model": model, "messages": messages, "temperature": temperature}
    try:
        r = requests.post(OPENROUTER_URL, headers=HEADERS, json=payload, timeout=180)
        r.raise_for_status()
        response = r.json()
        return response["choices"][0]["message"]["content"]
    except Exception as e:
        print(f"❌ OpenRouter custom call failed: {e}")
        return ""

def add_sop_metrics_via_llm(excel_file_path: str, model: str, temperature: float = 0):
    """Standalone: read the five statements from Excel, ask LLM to derive SOPs, and append a SOP_Metrics sheet.
    - excel_file_path: path to the existing Excel with five statements
    - model: OpenRouter model id to use (e.g., 'openai/gpt-4o', 'anthropic/claude-3-5-sonnet')
    - temperature: sampling temperature
    """
    print(f"🧠 Deriving SOP via LLM model: {model}")
    xl = pd.ExcelFile(excel_file_path)
    # Build a compact textual context from the five statements
    context_parts = []
    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(excel_file_path, sheet_name=sheet)
            if df is not None and not df.empty:
                sample = df.head(50)  # limit tokens
                context_parts.append(f"\n### {sheet}\n" + sample.to_csv(index=False))
        except Exception:
            continue
    context_text = "\n".join(context_parts)

    system_prompt = (
        "You are a senior financial analyst. From the provided five primary statements, "
        "derive a concise set of key SOP metrics strictly from your financial knowledge and the tables. "
        "Prioritize the latest Group/consolidated period. Output only a CSV with columns: Metric,Value,Source Term,Calculation Details. "
        "Use '-' if a value is explicitly dash/blank. Do not invent columns."
    )
    user_prompt = (
        "Five statements (CSV excerpts by sheet). Extract and compute metrics such as Revenue, Gross Profit, Operating Profit, "
        "Profit Before Tax, Profit for the Period, Total Assets, Total Liabilities, Total Equity, Cash and Cash Equivalents, "
        "Cash From Operating/Investing/Financing Activities, EPS, Book Value, and any other standard KPIs you judge essential.\n" 
        + context_text
    )

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    content = call_openrouter_custom_model(messages, model=model, temperature=temperature)
    if not content:
        print("❌ LLM returned empty content; skipping SOP write")
        return

    # Try to parse CSV-like response (robust handling of code fences and malformed rows)
    def _parse_llm_csv_to_kv(text: str) -> pd.DataFrame:
        # Strip markdown code fences like ```csv ... ``` or ``` ... ```
        fenced_blocks = re.findall(r"```[a-zA-Z]*\n([\s\S]*?)```", text)
        if fenced_blocks:
            text = fenced_blocks[0]
        # Remove a leading 'csv' line if present
        text = re.sub(r"^(csv|CSV)\s*\n", "", text.strip())
        # First attempt: pandas
        try:
            df = pd.read_csv(StringIO(text))
        except Exception:
            df = pd.DataFrame()
        # If pandas failed or wrong shape, try manual csv parsing
        if df.empty or len(df.columns) < 2:
            try:
                import csv as _csv
                rows = []
                for row in _csv.reader(StringIO(text)):
                    if not row:
                        continue
                    header_lower = [c.strip().lower() for c in row]
                    if "metric" in header_lower and "value" in header_lower:
                        # skip header row
                        continue
                    if len(row) >= 2:
                        rows.append([row[0], row[1]])
                if rows:
                    df = pd.DataFrame(rows, columns=["Metric", "Value"])
            except Exception:
                pass
        # Normalize columns to Metric/Value
        if not df.empty:
            lower_cols = {c.lower(): c for c in df.columns}
            if "metric" in lower_cols and "value" in lower_cols:
                df = df[[lower_cols["metric"], lower_cols["value"]]].rename(columns={lower_cols["metric"]: "Metric", lower_cols["value"]: "Value"})
            elif len(df.columns) >= 2:
                first, second = df.columns[:2]
                df = df[[first, second]].rename(columns={first: "Metric", second: "Value"})
        return df

    try:
        sop_df = _parse_llm_csv_to_kv(content)
        if sop_df is None or sop_df.empty:
            raise ValueError("Empty SOP df after parsing")
    except Exception:
        # Fallback: put raw content into a single key/value row to avoid single-cell mismatch
        sop_df = pd.DataFrame({"Metric": ["LLM_Output"], "Value": [content]})

    with pd.ExcelWriter(excel_file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        sop_df.to_excel(writer, sheet_name="SOP_Metrics", index=False)
        worksheet = writer.sheets["SOP_Metrics"]
        apply_excel_formatting(worksheet, sop_df)
    print("✅ LLM SOP_Metrics sheet written")

# =====================
# Step 1: Extract PDF with LLMWhisperer
# =====================
def get_text_cache_filename(pdf_path):
    """Generate cache filename based on PDF name"""
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    return os.path.join("Extracted_Text", f"{base_name}_extracted_text.txt")

def load_cached_text(pdf_path):
    """Load previously extracted text from cache file"""
    cache_file = get_text_cache_filename(pdf_path)
    
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                cached_text = f.read().strip()
            
            if cached_text:
                print(f"📋 Found cached text file: {cache_file}")
                print(f"✅ Loaded cached text! Length: {len(cached_text)} characters")
                return cached_text
            else:
                print(f"⚠️ Cached text file is empty: {cache_file}")
                return ""
        except Exception as e:
            print(f"❌ Error reading cached text: {e}")
            return ""
    else:
        print(f"📭 No cached text file found: {cache_file}")
        return ""

def save_extracted_text(pdf_path, text):
    """Save extracted text to cache file for future use"""
    if not text:
        return
    
    cache_file = get_text_cache_filename(pdf_path)
    try:
        # Ensure the Extracted_Text directory exists
        os.makedirs(os.path.dirname(cache_file), exist_ok=True)
        
        with open(cache_file, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"💾 Saved extracted text to: {cache_file}")
    except Exception as e:
        print(f"❌ Error saving text cache: {e}")

def create_sample_text_file(pdf_path):
    """Create a sample text file for testing (if you want to manually add extracted text)"""
    cache_file = get_text_cache_filename(pdf_path)
    
    if not os.path.exists(cache_file):
        sample_text = """# Sample Annual Report Text
# 
# Instructions: Replace this content with the actual extracted text from your PDF
# The AI will process this text to extract financial statements
#
# You can:
# 1. Copy-paste text from your PDF here
# 2. Use any PDF to text converter output
# 3. Let LLMWhisperer extract it automatically (when quota available)

SAMPLE CONTENT - REPLACE WITH ACTUAL ANNUAL REPORT TEXT
"""
        try:
            # Ensure the Extracted_Text directory exists
            os.makedirs(os.path.dirname(cache_file), exist_ok=True)
            
            with open(cache_file, 'w', encoding='utf-8') as f:
                f.write(sample_text)
            print(f"📝 Created sample text file: {cache_file}")
            print(f"💡 Edit this file with your PDF text content to skip LLMWhisperer")
        except Exception as e:
            print(f"❌ Error creating sample text file: {e}")
    else:
        print(f"📋 Text file already exists: {cache_file}")


def extract_with_pdfplumber_fallback(pdf_path):
    """Fallback extraction using pdfplumber if LLMWhisperer fails"""
    try:
        import pdfplumber
        
        print("🔄 Using pdfplumber as fallback...")
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            print(f"📄 Found {len(pdf.pages)} pages")
            
            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    text += f"\n<<<PAGE {page_num + 1}>>>\n"
                    text += page_text
                    
        return text
    except ImportError:
        print("❌ pdfplumber not installed. Run: pip install pdfplumber")
        return None
    except Exception as e:
        print(f"❌ pdfplumber extraction error: {e}")
        return None

def extract_pdf_text(pdf_path, wait_timeout=180, poll_interval=5):
    """Extract text from PDF using LLMWhisperer API (with caching and fallback)"""
    
    # First, try to load from cache
    cached_text = load_cached_text(pdf_path)
    if cached_text:
        return cached_text
    
    # If no cache, extract using LLMWhisperer with key rotation
    print(f"🔄 No cached text found. Extracting from PDF...")
    print(f"🔑 Key Status: {key_manager.get_key_status()}")
    
    whisper_job = None
    client = None
    
    # Try each available API key until one works or all fail
    while key_manager.has_available_keys():
        current_key = key_manager.get_current_key()
        if not current_key:
            break
        
        key_manager.total_attempts += 1
        print(f"🚀 Attempt {key_manager.total_attempts}: Trying LLMWhisperer extraction with API key {key_manager.current_key_index + 1}...")
        
        # Use LLMWhisperer client with current key
        client = LLMWhispererClientV2(api_key=current_key)
        try:
            whisper_job = client.whisper(
                file_path=pdf_path,
                wait_for_completion=True,
                wait_timeout=wait_timeout,
                output_mode="layout_preserving",
                mode="high_quality"
            )
            
            # If we get here, the API call succeeded
            key_manager.mark_key_successful()
            print(f"✅ LLMWhisperer API call successful with key {key_manager.current_key_index + 1}")
            break
            
        except LLMWhispererClientException as e:
            error_message = str(e).lower()
            print(f"❌ Error with API key {key_manager.current_key_index + 1}: {e}")
            
            # Check if this is a quota/limit error (more comprehensive detection)
            quota_keywords = [
                'quota', 'limit', 'exceeded', 'daily', 'usage', 'rate limit',
                'too many requests', '429', 'insufficient', 'exhausted',
                'maximum', 'threshold', 'billing', 'credit', 'subscription'
            ]
            
            if any(keyword in error_message for keyword in quota_keywords):
                if not key_manager.mark_key_failed("Daily/usage limit reached"):
                    print(f"⚠️ Falling back to pdfplumber extraction...")
                    break
                continue
            else:
                # Non-quota error (file issues, network, etc.)
                print(f"❌ Non-quota error encountered: {e}")
                print(f"⚠️ This might be a file or network issue, trying fallback...")
                break
        except Exception as e:
            # Catch any other unexpected errors
            print(f"❌ Unexpected error with API key {key_manager.current_key_index + 1}: {e}")
            
            if not key_manager.mark_key_failed("Unexpected error"):
                print(f"⚠️ All API keys exhausted, trying fallback...")
                break
            continue
    
    # If we have a successful whisper_job, process it
    if whisper_job is not None and client is not None:
        try:
            print(f"📊 Final Key Summary: {key_manager.get_summary()}")
            return process_whisper_job(whisper_job, client, wait_timeout, poll_interval, pdf_path)
        except Exception as e:
            print(f"❌ Error processing whisper job: {e}")
    
    # If all API keys failed or we had errors, try fallback
    print("⚠️ All LLMWhisperer API keys exhausted or failed. Trying pdfplumber fallback...")
    print(f"📊 Final Key Summary: {key_manager.get_summary()}")
    
    text = extract_with_pdfplumber_fallback(pdf_path)
    if text:
        save_extracted_text(pdf_path, text)
        print("✅ Fallback extraction successful")
    else:
        print("❌ Fallback extraction also failed")
    
    return text or ""

def process_whisper_job(whisper_job, client, wait_timeout, poll_interval, pdf_path):
    """Process a whisper job and return the extracted text"""
    
    if whisper_job.get("status") != "processed":
        whisper_hash = whisper_job.get("whisper_hash")
        if not whisper_hash:
            print("❌ No whisper hash received")
            return ""
            
        print(f"⏳ Waiting for processing to complete (hash: {whisper_hash})")
        status = whisper_job
        waited = 0
        while waited < wait_timeout:
            try:
                status = client.whisper_status(whisper_hash=whisper_hash)
                current_status = status.get("status")
                print(f"⏱️ Status after {waited}s: {current_status}")
                
                if current_status == "processed":
                    break
                elif current_status == "failed":
                    print("❌ LLMWhisperer processing failed")
                    return ""
                    
                time.sleep(poll_interval)
                waited += poll_interval
            except LLMWhispererClientException as e:
                print(f"❌ Error checking status: {e}")
                return ""
        
        if status.get("status") != "processed":
            print(f"⏰ Processing timed out after {wait_timeout}s")
            return ""
        
        try:
            print("📥 Retrieving processed text...")
            whisper_job = client.whisper_retrieve(whisper_hash=whisper_hash)
        except LLMWhispererClientException as e:
            print(f"❌ Error retrieving results: {e}")
            return ""

    extraction = whisper_job.get("extraction", {})
    result_text = extraction.get("result_text", "")
    
    if result_text:
        print(f"✅ Text extraction completed! Length: {len(result_text)} characters")
        save_extracted_text(pdf_path, result_text)
        return result_text
    else:
        print("❌ No text extracted from PDF")
        return ""

# =====================
# Step 2: Extraction Functions for Each Statement
# =====================
def extract_table(text, statement_name):
    system_prompt = f"""You are an expert financial data extractor specializing in annual reports.
    
    Your task: Find and extract the "{statement_name}" table from the provided text.
    
    PRIORITY RULES (CRITICAL - MUST FOLLOW):
    1. **QUARTERLY DATA MANDATORY**: ALWAYS choose quarterly data (Q1, Q2, Q3, Q4, or quarter-end dates like "31 Mar", "30 Jun", "30 Sep", "31 Dec") - NEVER use annual data
    2. **GROUP/CONSOLIDATED MANDATORY**: ALWAYS choose "Group" or "CONSOLIDATED" tables over "Company" tables
    3. **MOST RECENT QUARTER**: When multiple quarters are available, prioritize the most recent quarter
    4. **SKIP NON-QUARTERLY**: If no quarterly data is available, return "No quarterly data found" instead of annual data
    
    CRITICAL EXTRACTION REQUIREMENTS:
    1. **INCLUDE ALL ROW DESCRIPTIONS**: Always include the full description/name of each line item in the first column
    2. **PRESERVE ITEM HIERARCHY**: Maintain the structure of main items, sub-items, and totals exactly as shown
    3. **COMPLETE LINE ITEMS**: Don't abbreviate or truncate the descriptions of financial statement items
    4. **EXACT COLUMN HEADERS**: Use the original column titles from the document without modification
    5. **ALL NUMERICAL VALUES**: Include every number exactly as shown (including ('000) notation)
    6. **PRESERVE DASHES**: When a column has a dash (-), keep it as a dash - do NOT replace with values from other columns
    7. **MAINTAIN COLUMN ALIGNMENT**: Ensure each value stays in its correct column - dashes indicate no value for that period
    8. **PROPER FORMATTING**: Create a clean markdown table but preserve all original content and alignment
    
    WHAT TO INCLUDE:
    ✅ Full line item descriptions (e.g., "Revenue", "Cost of goods sold", "Gross profit", etc.)
    ✅ All numerical values in their original format
    ✅ Sub-totals and totals with proper hierarchy
    ✅ Notes and references attached to line items
    ✅ Original column headers with dates and units
    ✅ All rows including zeros and blank entries
    
    WHAT TO AVOID:
    ❌ Empty columns that are just separators (|, ---, spaces only)
    ❌ Abbreviating or shortening line item descriptions
    ❌ Modifying numerical values or formats
    ❌ Changing column header names
    ❌ Omitting any rows or data
    
    TABLE SELECTION LOGIC (MANDATORY):
    - ONLY search for tables with "Group" or "CONSOLIDATED" in the title
    - ONLY look for quarterly dates (Mar, Jun, Sep, Dec, Q1, Q2, Q3, Q4)
    - NEVER use annual data or Company tables
    - If no quarterly Group/CONSOLIDATED data found, return "No quarterly Group/CONSOLIDATED data found"
    
    EXAMPLE FORMAT:
    | Line Item Description | 2024 Rs.'000 | 2023 Rs.'000 |
    |----------------------|--------------|--------------|
    | Revenue              | 150,000      | 140,000      |
    | Cost of sales        | (90,000)     | (85,000)     |
    | Gross profit         | 60,000       | 55,000       |
    
    QUALITY CHECKS (MANDATORY):
    - Every row has a meaningful description in the first column
    - All numerical values are preserved with original formatting
    - Column headers match the source document exactly
    - Hierarchical structure is maintained (main items, sub-items, totals)
    - ONLY Group/CONSOLIDATED tables are used
    - ONLY quarterly data is used (no annual data)
    - If no quarterly Group/CONSOLIDATED data exists, return appropriate message
    
    Return ONLY the clean markdown table with complete descriptions and original headers. No explanations, no extra text."""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": text}
    ]
    return call_openrouter(messages)

def extract_profit_or_loss(text):
    # Try Statement of Profit or Loss first, then fallback to Statement of Income
    result = extract_table(text, "STATEMENT OF PROFIT OR LOSS")
    if not result or "No table found" in result or len(result.strip()) < 100:
        print("📋 Statement of Profit or Loss not found, trying Statement of Income...")
        result = extract_table(text, "STATEMENT OF INCOME")
    return result

def extract_comprehensive_income(text):
    return extract_table(text, "STATEMENT OF COMPREHENSIVE INCOME")

def extract_financial_position(text):
    return extract_table(text, "STATEMENT OF FINANCIAL POSITION")

def extract_changes_in_equity(text):
    """Special handling for STATEMENT OF CHANGES IN EQUITY with enhanced row descriptions"""
    system_prompt = f"""You are an expert financial data extractor specializing in annual reports.
    
    Your task: Find and extract the "STATEMENT OF CHANGES IN EQUITY" table from the provided text.
    
    CRITICAL REQUIREMENTS FOR CHANGES IN EQUITY:
    1. **INCLUDE ALL ROW DESCRIPTIONS**: This is MANDATORY for equity statements
    2. **COMPLETE EQUITY COMPONENT NAMES**: Include full names like:
       - "Stated Capital" or "Share Capital"
       - "ESOP Reserve" or "Employee Share Option Plan Reserve"
       - "General Reserve" or "Statutory Reserve"
       - "Retained Earnings" or "Accumulated Profits"
       - "Other Reserves" or "Fair Value Reserve"
       - "Total Equity"
    3. **MOVEMENT DESCRIPTIONS**: Include all movement types like:
       - "Balance at beginning of period"
       - "Issue of shares" or "Rights issue"
       - "Transfer to reserves"
       - "Dividend paid"
       - "Profit for the period"
       - "Other comprehensive income"
       - "Balance at end of period"
    
    PRIORITY RULES (CRITICAL - MUST FOLLOW):
    1. **QUARTERLY DATA MANDATORY**: ALWAYS choose quarterly data - NEVER use annual data
    2. **GROUP/CONSOLIDATED MANDATORY**: ALWAYS choose "Group" or "CONSOLIDATED" table over "Company" table
    3. **MOST RECENT QUARTER**: Prioritize the most recent quarter
    4. **SKIP NON-QUARTERLY**: If no quarterly data is available, return "No quarterly Group/CONSOLIDATED data found"
    
    FORMAT REQUIREMENTS:
    - First column MUST contain complete descriptions of equity components and movements
    - Include ALL numerical columns with original headers
    - Preserve exact column titles (dates, currency notations)
    - Maintain hierarchical structure of equity movements
    
    EXAMPLE FORMAT:
    | Equity Component/Movement | As at 31 Mar 2025 Rs.'000 | As at 31 Dec 2024 Rs.'000 |
    |---------------------------|----------------------------|----------------------------|
    | Stated Capital            | 2,500,000                  | 2,500,000                  |
    | Balance at beginning      | 2,500,000                  | 2,400,000                  |
    | Issue of shares           | -                          | 100,000                    |
    | Balance at end            | 2,500,000                  | 2,500,000                  |
    | ESOP Reserve              | 45,000                     | 40,000                     |
    | Retained Earnings         | 1,200,000                  | 1,100,000                  |
    | Total Equity              | 3,745,000                  | 3,640,000                  |
    
    QUALITY CHECKS (MANDATORY):
    - Every row has a meaningful, complete description
    - All equity components are clearly identified
    - All movements/transactions are properly described
    - No abbreviated or missing row descriptions
    - ONLY Group/CONSOLIDATED tables are used
    - ONLY quarterly data is used (no annual data)
    - If no quarterly Group/CONSOLIDATED data exists, return appropriate message
    
    Return ONLY the clean markdown table with complete equity descriptions. No explanations, no extra text."""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": text}
    ]
    return call_openrouter(messages)

def extract_cash_flows(text):
    """Enhanced cashflow extraction that can handle both quarterly and annual data, and both Group and Company tables"""
    system_prompt = f"""You are an expert financial data extractor specializing in annual reports.
    
    Your task: Find and extract the "STATEMENT OF CASH FLOWS" table from the provided text.
    
    PRIORITY RULES (in order of preference):
    1. **QUARTERLY DATA PREFERRED**: Look for quarterly data first (Q1, Q2, Q3, Q4, or quarter-end dates like "31 Mar", "30 Jun", "30 Sep", "31 Dec")
    2. **GROUP/CONSOLIDATED PREFERRED**: Look for "Group" or "CONSOLIDATED" tables first
    3. **FALLBACK TO ANNUAL**: If no quarterly data found, use annual data
    4. **FALLBACK TO COMPANY**: If no Group/CONSOLIDATED data found, use Company data
    5. **MOST RECENT PERIOD**: When multiple periods are available, prioritize the most recent period
    
    CRITICAL EXTRACTION REQUIREMENTS:
    1. **INCLUDE ALL ROW DESCRIPTIONS**: Always include the full description/name of each line item in the first column
    2. **PRESERVE ITEM HIERARCHY**: Maintain the structure of main items, sub-items, and totals exactly as shown
    3. **COMPLETE LINE ITEMS**: Don't abbreviate or truncate the descriptions of financial statement items
    4. **EXACT COLUMN HEADERS**: Use the original column titles from the document without modification
    5. **ALL NUMERICAL VALUES**: Include every number exactly as shown (including ('000) notation)
    6. **PRESERVE DASHES**: When a column has a dash (-), keep it as a dash - do NOT replace with values from other columns
    7. **MAINTAIN COLUMN ALIGNMENT**: Ensure each value stays in its correct column - dashes indicate no value for that period
    8. **PROPER FORMATTING**: Create a clean markdown table but preserve all original content and alignment
    
    WHAT TO INCLUDE:
    ✅ Full line item descriptions (e.g., "Cash from operating activities", "Cash from investing activities", etc.)
    ✅ All numerical values in their original format
    ✅ Sub-totals and totals with proper hierarchy
    ✅ Notes and references attached to line items
    ✅ Original column headers with dates and units
    ✅ All rows including zeros and blank entries
    
    WHAT TO AVOID:
    ❌ Empty columns that are just separators (|, ---, spaces only)
    ❌ Abbreviating or shortening line item descriptions
    ❌ Modifying numerical values or formats
    ❌ Changing column header names
    ❌ Omitting any rows or data
    
    TABLE SELECTION LOGIC (FLEXIBLE):
    - FIRST try to find tables with "Group" or "CONSOLIDATED" in the title with quarterly dates
    - THEN try to find tables with "Group" or "CONSOLIDATED" in the title with annual dates
    - THEN try to find tables with "Company" in the title with quarterly dates
    - FINALLY try to find tables with "Company" in the title with annual dates
    - If no cashflow table found at all, return "No cashflow data found"
    
    EXAMPLE FORMAT:
    | Line Item Description | 2024 Rs.'000 | 2023 Rs.'000 |
    |----------------------|--------------|--------------|
    | Cash from operating activities | 150,000 | 140,000 |
    | Cash from investing activities | (50,000) | (45,000) |
    | Cash from financing activities | (20,000) | (15,000) |
    | Net change in cash | 80,000 | 80,000 |
    
    QUALITY CHECKS (MANDATORY):
    - Every row has a meaningful description in the first column
    - All numerical values are preserved with original formatting
    - Column headers match the source document exactly
    - Hierarchical structure is maintained (main items, sub-items, totals)
    - If no cashflow data exists, return appropriate message
    
    Return ONLY the clean markdown table with complete descriptions and original headers. No explanations, no extra text."""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": text}
    ]
    return call_openrouter(messages)

# Explanatory notes extraction removed as per user request

# =====================
# Step 3: Save Multiple Sheets
# =====================
def is_ratio_or_percentage_column(col_name):
    """Check if a column contains ratios or percentages that shouldn't have zeros added"""
    col_lower = str(col_name).lower()
    ratio_indicators = [
        '%', 'percent', 'percentage', 'ratio', 'rate', 'times', 'x',
        'eps', 'earnings per share', 'p/e', 'price to earnings',
        'debt to equity', 'current ratio', 'quick ratio', 'roa', 'roe',
        'margin', 'yield', 'coverage', 'turnover', 'multiple'
    ]
    return any(indicator in col_lower for indicator in ratio_indicators)

# ('000) value conversion removed as per user request

def validate_column_alignment(df):
    """Validate and fix column alignment issues, ensuring dashes are preserved correctly"""
    if df.empty or len(df.columns) < 2:
        return df
    
    # Check for potential column misalignment issues
    for col_idx, col in enumerate(df.columns):
        if col_idx == 0:  # Skip the first column (usually row descriptions)
            continue
            
        col_values = df[col].astype(str).str.strip()
        
        # Check if this column has a high percentage of dashes (might indicate misalignment)
        dash_count = col_values.isin(['-', '--', '---']).sum()
        total_count = len(col_values)
        
        if total_count > 0 and dash_count / total_count > 0.8:
            print(f"⚠️ Warning: Column '{col}' has {dash_count}/{total_count} dashes - possible misalignment")
            
            # Check if the next column has values that might belong to this column
            if col_idx + 1 < len(df.columns):
                next_col = df.columns[col_idx + 1]
                next_col_values = df[next_col].astype(str).str.strip()
                next_col_numeric_count = sum(1 for val in next_col_values if val and val not in ['-', '--', '---', 'nan', 'NaN', 'None', ''] and val.replace(',', '').replace('.', '').replace('-', '').isdigit())
                
                if next_col_numeric_count > dash_count:
                    print(f"  🔧 Potential fix: Values from '{next_col}' might belong in '{col}'")
    
    return df

def clean_dataframe(df):
    """Clean the dataframe to remove NaN columns and improve data quality while preserving dashes"""
    # Remove completely empty columns (but be careful about columns with dashes)
    df = df.dropna(axis=1, how='all')
    
    # Remove columns that are just separators (contain only dashes, pipes, spaces)
    # BUT be more careful - only remove if ALL values are separators
    cols_to_drop = []
    for col in df.columns:
        col_values = df[col].astype(str).str.strip()
        # Only drop if ALL values are just separators (dashes, pipes, spaces) AND no meaningful data
        if col_values.str.match(r'^[-|\s]*$').all() and not any(val in ['-', '--', '---'] for val in col_values):
            cols_to_drop.append(col)
    df = df.drop(columns=cols_to_drop)
    
    # Clean column names
    df.columns = [str(col).strip().replace('Unnamed:', '') for col in df.columns]
    df.columns = [col for col in df.columns if not col.startswith('Unnamed')]
    
    # Remove rows that are completely empty or just separators
    # BUT preserve rows that have meaningful data even if some columns have dashes
    df = df[~df.astype(str).apply(lambda x: x.str.strip().str.match(r'^[-|\s]*$')).all(axis=1)]
    
    # Replace various forms of empty values with empty string, but PRESERVE dashes
    # Dashes should be kept as dashes, not replaced with empty strings
    df = df.replace(['nan', 'NaN', 'None'], '')
    # Note: We intentionally do NOT replace dashes here to preserve them in the data
    
    # Validate column alignment
    df = validate_column_alignment(df)
    
    return df

def markdown_to_df(md_table):
    try:
        # First, clean the markdown table text
        lines = md_table.strip().split('\n')
        cleaned_lines = []
        
        for line in lines:
            # Skip separator lines that are just dashes and pipes
            if line.strip() and not line.strip().replace('-', '').replace('|', '').replace(' ', ''):
                continue
            cleaned_lines.append(line)
        
        cleaned_md = '\n'.join(cleaned_lines)
        
        # Use more robust parsing to preserve column alignment
        df = pd.read_table(StringIO(cleaned_md), sep="|", engine="python", header=0, skipinitialspace=True)
        
        # Clean column names but preserve structure
        df.columns = [c.strip() for c in df.columns]
        
        # Remove unnamed/empty columns but be careful not to remove columns with dashes
        cols_to_keep = []
        for col in df.columns:
            if not col.startswith('Unnamed') and col != "":
                cols_to_keep.append(col)
        
        if cols_to_keep:
            df = df[cols_to_keep]
        
        # Clean each column but PRESERVE dashes
        for col in df.columns:
            # Convert to string and strip whitespace, but preserve dashes
            df[col] = df[col].astype(str).str.strip()
            # Ensure dashes are preserved as dashes, not converted to NaN
            df[col] = df[col].replace('nan', '-')
            df[col] = df[col].replace('NaN', '-')
            df[col] = df[col].replace('None', '-')
        
        # Apply additional cleaning but preserve dashes
        df = clean_dataframe(df)
        
        return df
    except Exception as e:
        print("⚠️ Could not parse markdown to DataFrame:", e)
        print("Raw markdown content:", md_table[:500] + "..." if len(md_table) > 500 else md_table)
        return pd.DataFrame({"data": [md_table]})

def apply_excel_formatting(worksheet, df):
    """Apply bold yellow formatting to headers and adjust column widths"""
    # Define styles
    header_font = Font(bold=True, color="000000")  # Black text
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply formatting to header row
    for col_num, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width

def save_all_to_excel(tables_dict, filename="PrimaryStatements.xlsx", sop_metrics=None, sop_source_terms=None, sop_calculations=None):
    """Save all tables to Excel with proper formatting"""
    # Ensure the Excel_Statements directory exists
    os.makedirs("Excel_Statements", exist_ok=True)
    
    # If filename doesn't include the folder path, add it
    if not filename.startswith("Excel_Statements"):
        filename = os.path.join("Excel_Statements", filename)
    
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Save financial statement tables
        for sheet, content in tables_dict.items():
            sheet_name = sheet[:30]  # Excel sheet name limit
            
            # Process financial statement tables
            df = markdown_to_df(content)
            
            if not df.empty and len(df.columns) > 0:
                # Save to Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet and apply formatting
                worksheet = writer.sheets[sheet_name]
                apply_excel_formatting(worksheet, df)
            else:
                # Fallback for problematic tables
                fallback_df = pd.DataFrame({"Raw Data": [content]})
                fallback_df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                apply_excel_formatting(worksheet, fallback_df)
        
        # SOP metrics sheet removed from the main save flow
    
    print(f"✅ Excel file saved as: {filename}")
    return filename

# =====================
# Step 4: SOP (Standard Operating Procedures) Calculations
# =====================

def clean_numeric_value(value):
    """Clean and convert a value to numeric, handling various formats"""
    if pd.isna(value) or value == '' or value == 'nan':
        return 0.0
    
    # Convert to string and clean
    str_value = str(value).strip()
    
    # Skip percentage values (contain %)
    if '%' in str_value:
        return 0.0
    
    # Handle dashes - return a special value to indicate dash should be preserved
    if str_value in ['-', '--', '---']:
        return 'DASH'  # Special marker to indicate dash should be preserved
    
    # Remove asterisks and other formatting
    str_value = str_value.replace('*', '').replace(',', '')
    
    # Handle negative values in parentheses
    if str_value.startswith('(') and str_value.endswith(')'):
        str_value = '-' + str_value[1:-1]
    
    # Extract only numbers and decimal points and minus signs
    str_value = re.sub(r'[^\d.-]', '', str_value)
    
    # Remove multiple decimal points (keep only the first one)
    parts = str_value.split('.')
    if len(parts) > 2:
        str_value = parts[0] + '.' + ''.join(parts[1:])
    
    try:
        return float(str_value)
    except:
        return 0.0

def to_number(value, default=0.0):
    """Convert a value to float safely. Returns default for non-numeric/dash values."""
    if value is None:
        return default
    # Treat common dash/empty markers as default
    if isinstance(value, str) and value.strip() in ['', '-', '--', '---', 'DASH', 'n/a', 'N/A', 'nil', 'Nil']:
        return default
    try:
        # If already numeric
        if isinstance(value, (int, float)):
            return float(value)
        # Clean string and parse
        s = str(value).replace(',', '').replace('*', '').strip()
        if s.startswith('(') and s.endswith(')'):
            s = '-' + s[1:-1]
        return float(s)
    except Exception:
        return default

def fmt_value(value):
    """Format a metric value for display. Numeric -> 2dp, otherwise string as-is."""
    if isinstance(value, (int, float)):
        return f"{float(value):,.2f}"
    # Try to coerce
    coerced = to_number(value, None)
    if coerced is not None:
        return f"{coerced:,.2f}"
    return str(value)

def find_value_in_dataframes_intelligent(search_terms, dataframes, exact_match=False):
    """Enhanced intelligent value finder using AI knowledge and semantic understanding
    CRITICAL: ALWAYS prioritizes quarterly data from Group/CONSOLIDATED tables
    Returns: (value, source_term_found)"""
    
    best_value = 0.0
    best_priority = -1
    best_source_term = ""
    
    for df in dataframes:
        if df.empty or len(df.columns) < 2:
            continue
        
        # Check column headers to determine priority
        # CRITICAL PRIORITY: Quarter Group > Quarter Company > Annual Group > Annual Company
        priority_score = 0
        column_headers = ' '.join([str(col).lower() for col in df.columns])
        
        # Check for quarterly data (PREFERRED - highest priority)
        quarter_indicators = ['mar', 'jun', 'sep', 'dec', 'q1', 'q2', 'q3', 'q4', '31 mar', '30 jun', '30 sep', '31 dec']
        has_quarterly_data = any(indicator in column_headers for indicator in quarter_indicators)
        if has_quarterly_data:
            priority_score += 100  # Much higher priority for quarterly data
        else:
            # For cashflow metrics, also allow annual data as fallback
            annual_indicators = ['2024', '2023', '2022', '2021', '2020', 'year', 'annual']
            has_annual_data = any(indicator in column_headers for indicator in annual_indicators)
            if has_annual_data:
                priority_score += 50  # Lower priority for annual data
            else:
                # Skip if neither quarterly nor annual data found
                continue
        
        # Check for Group vs Company (Group has higher priority)
        if 'group' in column_headers or 'consolidated' in column_headers:
            priority_score += 50  # High priority for Group/CONSOLIDATED
        elif 'company' in column_headers:
            priority_score += 10  # Lower priority for Company
        else:
            # For cashflow metrics, also allow data without explicit Group/Company designation
            priority_score += 5  # Very low priority for unidentified data
        
        # Find the rightmost column (most recent period) - higher priority
        rightmost_col_index = len(df.columns) - 1
        priority_score += rightmost_col_index
        
        first_col = df.iloc[:, 0].astype(str)
        
        for row_idx, row_text in enumerate(first_col):
            if pd.isna(row_text) or not row_text.strip():
                continue
            
            row_text_clean = str(row_text).strip()
            row_text_lower = row_text_clean.lower()
            
            # Calculate semantic similarity for each search term
            best_similarity = 0.0
            best_matching_term = ""
            for term in search_terms:
                term_lower = term.lower()
                # Exact match gets highest score
                if term_lower == row_text_lower:
                    similarity = 1.0
                    best_matching_term = term
                    break
                # Target contains term
                if term_lower in row_text_lower:
                    similarity = 0.9
                else:
                    # Word-by-word similarity
                    term_words = set(term_lower.split())
                    target_words = set(row_text_lower.split())
                    if term_words and target_words:
                        common_words = term_words.intersection(target_words)
                        similarity = len(common_words) / max(len(term_words), len(target_words))
                    else:
                        similarity = 0.0
                # Boost for financial synonyms
                similarity = boost_financial_similarity(term_lower, row_text_lower, similarity)
                if similarity > best_similarity:
                    best_similarity = similarity
                    best_matching_term = term
            
            # If we found a good match (similarity > 0.6), extract the value
            if best_similarity > 0.6:
                matched_row = df.iloc[row_idx]
                
                # Get values from columns - ALWAYS prioritize the LATEST/MOST RECENT column
                best_col_value = 0.0
                best_col_priority = -1
                
                for i in range(1, len(matched_row)):
                    col_header = str(df.columns[i]).lower()
                    
                    # Skip percentage, change, and ratio columns
                    if any(skip_word in col_header for skip_word in ['%', 'change', 'ratio', 'growth']):
                        continue
                    
                    val = clean_numeric_value(matched_row.iloc[i])
                    
                    # Handle dash marker - if we find a dash, preserve it and don't look further
                    if val == 'DASH':
                        best_col_value = '-'  # Return the actual dash
                        best_col_priority = float('inf')  # Highest priority to preserve dash
                        break  # Stop looking at other columns when we find a dash
                    
                    if val != 0 or str(matched_row.iloc[i]).strip() in ['0', '0.0']:
                        # Determine column priority based on dates
                        col_priority = 0
                        
                        # Higher priority for more recent years
                        if '2025' in col_header:
                            col_priority += 1000
                        elif '2024' in col_header:
                            col_priority += 500
                        elif '2023' in col_header:
                            col_priority += 100
                        
                        # Higher priority for more recent quarters/months
                        if any(recent in col_header for recent in ['mar 2025', 'q1 2025', '31 mar 2025']):
                            col_priority += 2000
                        elif any(recent in col_header for recent in ['dec 2024', 'q4 2024', '31 dec 2024']):
                            col_priority += 1500
                        
                        # Use leftmost position as tiebreaker (usually more recent)
                        col_priority += (100 - i)  # Earlier columns get higher priority
                        
                        if col_priority > best_col_priority:
                            best_col_value = val
                            best_col_priority = col_priority
                
                if best_col_value != 0 or best_col_value == '-':
                    current_priority = priority_score + best_col_priority + (best_similarity * 100)
                    
                    if current_priority > best_priority:
                        best_value = best_col_value
                        best_priority = current_priority
                        best_source_term = row_text_clean
    
    return best_value, best_source_term

def boost_financial_similarity(term, target, base_similarity):
    """Boost similarity for known financial term variations using AI knowledge"""
    
    financial_synonyms = {
        'revenue': ['income', 'earnings', 'sales', 'turnover', 'proceeds', 'total revenue', 'gross revenue'],
        'profit': ['earnings', 'income', 'surplus'],
        'net profit': ['net earnings', 'net income', 'profit for the period', 'profit for the year', 'profit/(loss) for the period'],
        'gross profit': ['gross earnings', 'gross income', 'net operating income'],
        'operating profit': ['operating income', 'operating earnings', 'ebit', 'operating profit before tax'],
        'profit before tax': ['earnings before tax', 'pre-tax profit', 'profit before taxation', 'profit/ (loss) before income tax'],
        'taxation': ['tax expense', 'income tax', 'tax provision', 'income tax expense'],
        'total assets': ['assets total', 'total asset'],
        'total liabilities': ['liabilities total', 'total liability'],
        'cash': ['cash and cash equivalents', 'cash equivalents'],
        'share price': ['price per share', 'market price', 'stock price', 'last traded', 'market price of ordinary share'],
        'earnings per share': ['eps', 'earning per share', 'basic earnings per share'],
        'depreciation': ['depreciation expense', 'depreciation cost'],
        'interest income': ['financing income', 'finance income'],
        'fee and commission income': ['net fee and commission income'],
        'net trading income': ['trading income'],
        'other operating income': ['net other operating income'],
        'total number of issued shares': ['number of shares', 'ordinary shares', 'number of ordinary shares'],
        'trade receivables': ['accounts receivable', 'debtors', 'customer receivables'],
        'trade payables': ['accounts payable', 'creditors', 'suppliers payable'],
        'inventories': ['stock', 'inventory', 'raw materials', 'finished goods'],
        'fixed assets': ['property, plant and equipment', 'ppe', 'tangible assets'],
        'current assets': ['total current assets'],
        'current liabilities': ['total current liabilities'],
        'total debt': ['total borrowings', 'debt', 'borrowings'],
        'book value': ['net asset value per share', 'net worth per share']
    }
    
    # Check if term has known synonyms in target
    for main_term, synonyms in financial_synonyms.items():
        if main_term in term:
            for synonym in synonyms:
                if synonym in target:
                    return max(base_similarity, 0.8)
    
    return base_similarity

def calculate_revenue_from_components(dataframes):
    """Calculate Revenue from income components using AI knowledge"""
    
    print("  🔧 Calculating Revenue from components:")
    
    # Components to ADD (positive income)
    add_components = [
        ['financing income'],
        ['net fee and commission income'], 
        ['net trading income'],
        ['net other operating income']
    ]
    
    # Components that could be gains OR losses (check sign)
    gain_loss_components = [
        ['net gains / (losses) from financial investments at fair value through profit or loss', 'net gains from financial investments'],
        ['net gains / (losses) from derecognition of financial assets', 'net gains from derecognition']
    ]
    
    total_revenue = 0.0
    found_components = []
    
    # Add the positive income components
    for component_terms in add_components:
        value, source_term = find_value_in_dataframes_intelligent(component_terms, dataframes)
        if value != 0:
            total_revenue += value
            found_components.append(f"{component_terms[0]}: +{value:,.0f}")
            print(f"    + {component_terms[0]}: +{value:,.2f} (from: {source_term})")
        else:
            print(f"    - {component_terms[0]}: Not found")
    
    # Handle gain/loss components (check if they're gains or losses)
    for component_terms in gain_loss_components:
        value, source_term = find_value_in_dataframes_intelligent(component_terms, dataframes)
        if value != 0:
            # If it's a negative value (loss), subtract it from revenue
            # If it's positive (gain), add it to revenue
            if value < 0:
                total_revenue += value  # Adding a negative number = subtraction
                found_components.append(f"{component_terms[0]}: {value:,.0f} (LOSS)")
                print(f"    - {component_terms[0]}: {value:,.2f} (LOSS - from: {source_term})")
            else:
                total_revenue += value
                found_components.append(f"{component_terms[0]}: +{value:,.0f} (GAIN)")
                print(f"    + {component_terms[0]}: +{value:,.2f} (GAIN - from: {source_term})")
        else:
            print(f"    - {component_terms[0]}: Not found")
    
    return total_revenue, found_components

def find_value_in_dataframes(search_terms, dataframes, exact_match=False):
    """Legacy function - kept for backward compatibility"""
    return find_value_in_dataframes_intelligent(search_terms, dataframes, exact_match)

def calculate_revenue_components(dataframes):
    """Calculate Revenue from income components. Handle gains/losses correctly:
    ADD: Financing Income + Net Fee and Commission Income + Net Trading Income + Net Other Operating Income
    SUBTRACT: Net Losses from Financial Investments and Derecognition (if they are losses in brackets)"""
    
    print("  🔧 Calculating Revenue from components:")
    
    # Components to ADD (positive income)
    add_components = [
        ['financing income'],
        ['net fee and commission income'], 
        ['net trading income'],
        ['net other operating income']
    ]
    
    # Components that could be gains OR losses (check sign)
    gain_loss_components = [
        ['net gains / (losses) from financial investments at fair value through profit or loss', 'net gains from financial investments'],
        ['net gains / (losses) from derecognition of financial assets', 'net gains from derecognition']
    ]
    
    total_revenue = 0.0
    found_components = []
    
    # Add the positive income components
    for component_terms in add_components:
        value, source_term = find_value_in_dataframes(component_terms, dataframes)
        if value != 0:
            total_revenue += value
            found_components.append(f"{component_terms[0]}: +{value:,.0f}")
            print(f"    + {component_terms[0]}: +{value:,.2f} (from: {source_term})")
        else:
            print(f"    - {component_terms[0]}: Not found")
    
    # Handle gain/loss components (check if they're gains or losses)
    for component_terms in gain_loss_components:
        value, source_term = find_value_in_dataframes(component_terms, dataframes)
        if value != 0:
            # If it's a negative value (loss), subtract it from revenue
            # If it's positive (gain), add it to revenue
            if value < 0:
                total_revenue += value  # Adding a negative number = subtraction
                found_components.append(f"{component_terms[0]}: {value:,.0f} (LOSS)")
                print(f"    - {component_terms[0]}: {value:,.2f} (LOSS - from: {source_term})")
            else:
                total_revenue += value
                found_components.append(f"{component_terms[0]}: +{value:,.0f} (GAIN)")
                print(f"    + {component_terms[0]}: +{value:,.2f} (GAIN - from: {source_term})")
        else:
            print(f"    - {component_terms[0]}: Not found")
    
    return total_revenue, found_components

def calculate_net_change_in_cash(dataframes):
    """Calculate Net Change in Cash"""
    operating, _ = find_value_in_dataframes(['net cash from operating activities', 'cash from operating'], dataframes)
    investing, _ = find_value_in_dataframes(['net cash from investing activities', 'cash from investing'], dataframes)
    financing, _ = find_value_in_dataframes(['net cash from financing activities', 'cash from financing'], dataframes)
    
    return operating + investing + financing

def calculate_enterprise_value(dataframes, market_cap=0):
    """Calculate Enterprise Value (Market Cap + Total Debt - Cash)"""
    total_debt, _ = find_value_in_dataframes(['total debt', 'total borrowings', 'debt'], dataframes)
    cash, _ = find_value_in_dataframes(['cash and cash equivalents', 'cash'], dataframes)
    
    return market_cap + total_debt - cash

def calculate_tax_rate(dataframes):
    """Calculate Tax Rate ((Total Tax Paid/Total Income)*100)"""
    tax_paid, _ = find_value_in_dataframes(['tax expense', 'income tax', 'taxation'], dataframes)
    total_income, _ = find_value_in_dataframes(['total income', 'gross income', 'revenue'], dataframes)
    
    if total_income != 0:
        return (tax_paid / total_income) * 100
    return 0.0

def calculate_effective_tax_rate(dataframes):
    """Calculate Effective Tax Rate ((Income Tax Expense/EBT)*100)"""
    tax_expense, _ = find_value_in_dataframes(['income tax expense', 'tax expense', 'taxation'], dataframes)
    ebt, _ = find_value_in_dataframes(['profit before tax', 'earnings before tax', 'profit before taxation'], dataframes)
    
    if ebt != 0:
        return (tax_expense / ebt) * 100
    return 0.0

def calculate_capital_expenditure(dataframes):
    """Calculate Capital Expenditure (Acquisition of PPE + Intangible Assets)"""
    ppe_acquisition, _ = find_value_in_dataframes(['acquisition of property', 'purchase of property', 'additions to property'], dataframes)
    intangible_acquisition, _ = find_value_in_dataframes(['acquisition of intangible', 'purchase of intangible', 'additions to intangible'], dataframes)
    
    return abs(ppe_acquisition) + abs(intangible_acquisition)  # Usually negative in cash flow, so take absolute

def extract_sop_metrics_knowledge_based(tables_dict, extracted_text=""):
    """Extract SOP metrics using AI knowledge and understanding of financial statements
    This approach uses comprehensive knowledge of financial statement patterns and terminology.
    Accepts a mapping of sheet_name -> markdown table string (legacy) and converts to DataFrames."""
    
    # Convert markdown tables to dataframes (legacy path)
    dataframes = []
    for sheet_name, content in tables_dict.items():
        df = markdown_to_df(content)
        if not df.empty:
            dataframes.append(df)
    
    return extract_sop_metrics_from_dataframes_knowledge_based(dataframes, extracted_text)


def extract_sop_metrics_from_dataframes_knowledge_based(dataframes, extracted_text=""):
    """Extract SOP metrics directly from already-loaded pandas DataFrames using AI knowledge.
    ALWAYS prioritizes quarterly data from Group/CONSOLIDATED tables."""
    
    print("🧠 Starting knowledge-based SOP metrics extraction with quarterly Group/CONSOLIDATED priority...")
    
    # Initialize SOP metrics dictionary with tracking
    sop_metrics = {}
    sop_source_terms = {}  # Track what term was actually found
    sop_calculations = {}  # Track calculation details
    
    # =====================
    # COMPREHENSIVE FINANCIAL METRICS EXTRACTION
    # =====================
    print("📊 Extracting comprehensive financial metrics using AI knowledge...")
    print("🎯 PRIORITY: Quarterly data from Group/CONSOLIDATED tables only")
    
    # Define comprehensive search patterns based on financial knowledge
    financial_metrics = {
        # REVENUE AND INCOME METRICS
        'Revenues': {
            'search_terms': [
                'revenue', 'total revenue', 'income', 'total income', 'sales', 'turnover',
                'interest income', 'fee income', 'commission income', 'trading income',
                'net fee and commission income', 'net trading income', 'net other operating income',
                'financing income', 'revenue from contracts with customers'
            ],
            'calculation_type': 'direct_or_calculated'
        },
        'Gross profit': {
            'search_terms': [
                'gross profit', 'gross income', 'net operating income', 'operating income',
                'gross earnings', 'gross surplus'
            ],
            'calculation_type': 'direct'
        },
        'Operating Profits': {
            'search_terms': [
                'operating profit', 'operating income', 'operating earnings', 'ebit',
                'operating profit before tax', 'operating profit before vat',
                'operating profit before vat on financial services and social security contribution levy'
            ],
            'calculation_type': 'direct'
        },
        'Net Profit': {
            'search_terms': [
                'net profit', 'profit for the period', 'profit for the year', 'net income',
                'profit/(loss) for the period', 'net earnings', 'profit after tax'
            ],
            'calculation_type': 'direct'
        },
        'Profit Before Tax': {
            'search_terms': [
                'profit before tax', 'earnings before tax', 'profit before taxation',
                'profit/ (loss) before income tax', 'pre-tax profit', 'ebt'
            ],
            'calculation_type': 'direct'
        },
        
        # TAXATION METRICS
        'Taxation': {
            'search_terms': [
                'taxation', 'tax expense', 'income tax', 'income tax expense', 'tax provision',
                'current tax', 'deferred tax', 'total tax expense'
            ],
            'calculation_type': 'direct'
        },
        
        # INTEREST METRICS
        'Interest Income': {
            'search_terms': [
                'interest income', 'financing income', 'finance income', 'interest earned'
            ],
            'calculation_type': 'direct'
        },
        'Interest Expense': {
            'search_terms': [
                'interest expense', 'interest paid', 'finance costs', 'interest cost'
            ],
            'calculation_type': 'direct'
        },
        
        # BALANCE SHEET METRICS
        'Fixed Assets': {
            'search_terms': [
                'property, plant, equipment and right-of-use assets', 'fixed assets',
                'property, plant and equipment', 'ppe', 'tangible assets'
            ],
            'calculation_type': 'direct'
        },
        'Inventory': {
            'search_terms': [
                'inventories', 'stock', 'inventory', 'raw materials', 'finished goods'
            ],
            'calculation_type': 'direct'
        },
        'Trade Receivables': {
            'search_terms': [
                'trade receivables', 'accounts receivable', 'debtors',
                'financial assets at amortized cost - financing and receivables to other customers',
                'loans and advances to customers', 'customer receivables'
            ],
            'calculation_type': 'direct'
        },
        'Cash': {
            'search_terms': [
                'cash', 'cash and cash equivalents', 'cash equivalents', 'cash and bank balances'
            ],
            'calculation_type': 'direct'
        },
        'Current Assets': {
            'search_terms': [
                'current assets', 'total current assets'
            ],
            'calculation_type': 'direct'
        },
        'Total Assets': {
            'search_terms': [
                'total assets', 'assets total', 'total asset'
            ],
            'calculation_type': 'direct'
        },
        'Total Equity': {
            'search_terms': [
                'total equity', 'equity total', 'shareholders equity', 'stockholders equity'
            ],
            'calculation_type': 'direct'
        },
        'Trade Payables': {
            'search_terms': [
                'trade payables', 'accounts payable', 'creditors', 'suppliers payable'
            ],
            'calculation_type': 'direct'
        },
        'Current Liabilities': {
            'search_terms': [
                'current liabilities', 'total current liabilities'
            ],
            'calculation_type': 'direct'
        },
        'Total Liabilities': {
            'search_terms': [
                'total liabilities', 'liabilities total', 'total liability'
            ],
            'calculation_type': 'direct'
        },
        'Total Debt': {
            'search_terms': [
                'total debt', 'total borrowings', 'debt', 'borrowings',
                'short term borrowings', 'long term borrowings', 'bank borrowings'
            ],
            'calculation_type': 'direct'
        },
        
        # CASH FLOW METRICS (QUARTERLY)
        'OCF Qtrly': {
            'search_terms': [
                'net cash from / (used in) operating activities', 'cash from operating activities',
                'operating cash flow', 'net cash from operating'
            ],
            'calculation_type': 'direct'
        },
        'ICF Qtrly': {
            'search_terms': [
                'net cash from / (used in) investing activities', 'cash from investing activities',
                'investing cash flow', 'net cash from investing'
            ],
            'calculation_type': 'direct'
        },
        'FCF Qtrly': {
            'search_terms': [
                'net cash from / (used in) financing activities', 'cash from financing activities',
                'financing cash flow', 'net cash from financing'
            ],
            'calculation_type': 'direct'
        },
        
        # DEPRECIATION AND AMORTIZATION (QUARTERLY)
        'Depreciation Qtrly': {
            'search_terms': [
                'depreciation and amortization of property, plant, equipment and right-of-use assets',
                'depreciation of property, plant and equipment', 'depreciation expense',
                'depreciation', 'depreciation and amortization'
            ],
            'calculation_type': 'direct'
        },
        'Amortization Qtrly': {
            'search_terms': [
                'amortization', 'amortization expense', 'amortization of intangible assets'
            ],
            'calculation_type': 'direct'
        },
        
        # CAPITAL EXPENDITURE (QUARTERLY)
        'Capital Exp Qtrly': {
            'search_terms': [
                'acquisition of property, plant & equipment', 'acquisition of intangible assets',
                'purchase of property, plant and equipment', 'additions to property, plant and equipment',
                'capital expenditure', 'capex'
            ],
            'calculation_type': 'direct'
        },
        
        # NET BORROWINGS (QUARTERLY)
        'Net Borrowings Qrtly': {
            'search_terms': [
                'net borrowings', 'short term borrowings', 'long term borrowings',
                'principal element of lease payment', 'lease payments'
            ],
            'calculation_type': 'direct'
        },
        
        # SHARE AND MARKET METRICS
        'Share Price Quaterly': {
            'search_terms': [
                'share price', 'price per share', 'market price', 'last traded',
                'market price of ordinary share', 'stock price', 'last traded price'
            ],
            'calculation_type': 'direct'
        },
        'Tot. No. of Shares': {
            'search_terms': [
                'total number of issued shares', 'number of shares', 'ordinary shares',
                'number of ordinary shares', 'issued shares', 'total shares'
            ],
            'calculation_type': 'direct'
        },
        'EPS Actuals': {
            'search_terms': [
                'earnings per share', 'eps', 'basic earnings per share', 'earning per share',
                'earnings per share - basic / diluted (in rs.)', 'basic eps'
            ],
            'calculation_type': 'direct'
        },
        'Book Value': {
            'search_terms': [
                'book value', 'net asset value per share', 'net asset value per ordinary share',
                'book value per share', 'net worth per share'
            ],
            'calculation_type': 'direct'
        }
    }
    
    # Extract all direct metrics using intelligent search
    for metric_name, config in financial_metrics.items():
        print(f"🔍 Searching for {metric_name}...")
        
        if config['calculation_type'] == 'direct':
            value, source_term = find_value_in_dataframes_intelligent(config['search_terms'], dataframes)
            sop_metrics[metric_name] = value
            sop_source_terms[metric_name] = source_term if source_term else "Not found"
            sop_calculations[metric_name] = "Direct extraction using AI knowledge"
            
            if to_number(value, 0.0) != 0.0 or (isinstance(value, str) and value.strip() in ['-', '--', '---']):
                print(f"  ✅ {metric_name}: {fmt_value(value)} (from: {source_term})")
            else:
                print(f"  ❌ {metric_name}: Not found")
        
        elif config['calculation_type'] == 'direct_or_calculated':
            # Try direct extraction first
            value, source_term = find_value_in_dataframes_intelligent(config['search_terms'], dataframes)
            
            if to_number(value, 0.0) == 0.0 and not (isinstance(value, str) and value.strip() in ['-', '--', '---']):
                # If not found directly, try to calculate from components
                print(f"  🔧 {metric_name} not found directly, attempting calculation...")
                if metric_name == 'Revenues':
                    value, revenue_components = calculate_revenue_from_components(dataframes)
                    sop_calculations[metric_name] = " + ".join(revenue_components) if revenue_components else "No components found"
                    source_term = "Calculated from components"
                else:
                    sop_calculations[metric_name] = "Direct extraction attempted, not found"
                    source_term = "Not found"
            
            sop_metrics[metric_name] = value
            sop_source_terms[metric_name] = source_term
            
            if to_number(value, 0.0) != 0.0 or (isinstance(value, str) and value.strip() in ['-', '--', '---']):
                print(f"  ✅ {metric_name}: {fmt_value(value)} ({source_term})")
        else:
                print(f"  ❌ {metric_name}: Not found")
    
    # =====================
    # CALCULATED METRICS USING FINANCIAL KNOWLEDGE
    # =====================
    print("\n🧮 Calculating derived metrics using financial knowledge...")
    
    # Market Capitalization = Share Price × Total Number of Issued Shares
    share_price = to_number(sop_metrics.get('Share Price Quaterly', 0), 0.0)
    total_shares = to_number(sop_metrics.get('Tot. No. of Shares', 0), 0.0)
    if share_price > 0 and total_shares > 0:
        market_cap = share_price * total_shares
        sop_metrics['Market Capitalization'] = market_cap
        sop_source_terms['Market Capitalization'] = "Calculated from share price and shares"
        sop_calculations['Market Capitalization'] = f"{share_price:,.2f} × {total_shares:,.0f} = {market_cap:,.2f}"
        print(f"  ✅ Market Capitalization: {market_cap:,.2f}")
    
    # Net Change in Cash = Operating + Investing + Financing (Quarterly)
    operating_cash = to_number(sop_metrics.get('OCF Qtrly', 0), 0.0)
    investing_cash = to_number(sop_metrics.get('ICF Qtrly', 0), 0.0)
    financing_cash = to_number(sop_metrics.get('FCF Qtrly', 0), 0.0)
    net_change_cash = operating_cash + investing_cash + financing_cash
    sop_metrics['Net Change In Cash Qtrly'] = net_change_cash
    sop_source_terms['Net Change In Cash Qtrly'] = "Calculated from quarterly cash flows"
    sop_calculations['Net Change In Cash Qtrly'] = f"{operating_cash:,.2f} + {investing_cash:,.2f} + {financing_cash:,.2f} = {net_change_cash:,.2f}"
    print(f"  ✅ Net Change In Cash Qtrly: {net_change_cash:,.2f}")
    
    # Enterprise Value = Market Cap + Total Debt - Cash
    market_cap = to_number(sop_metrics.get('Market Capitalization', 0), 0.0)
    total_debt = to_number(sop_metrics.get('Total Debt', 0), 0.0)
    cash = to_number(sop_metrics.get('Cash', 0), 0.0)
    enterprise_value = market_cap + total_debt - cash
    sop_metrics['Enterprise Value EV'] = enterprise_value
    sop_source_terms['Enterprise Value EV'] = "Calculated from market cap, debt, and cash"
    sop_calculations['Enterprise Value EV'] = f"{market_cap:,.2f} + {total_debt:,.2f} - {cash:,.2f} = {enterprise_value:,.2f}"
    print(f"  ✅ Enterprise Value: {enterprise_value:,.2f}")
    
    # Tax Rate = (Taxation / Revenue) × 100
    taxation = to_number(sop_metrics.get('Taxation', 0), 0.0)
    revenue = to_number(sop_metrics.get('Revenues', 0), 0.0)
    if revenue != 0:
        tax_rate = (abs(taxation) / revenue) * 100
        sop_metrics['Tax Rate'] = tax_rate
        sop_source_terms['Tax Rate'] = "Calculated from taxation and revenue"
        sop_calculations['Tax Rate'] = f"({abs(taxation):,.2f} / {revenue:,.2f}) × 100 = {tax_rate:.2f}%"
        print(f"  ✅ Tax Rate: {tax_rate:.2f}%")
    else:
        tax_rate = 0.0
        sop_metrics['Tax Rate'] = tax_rate
    
    # Effective Tax Rate = (Taxation / Profit Before Tax) × 100
    profit_before_tax = to_number(sop_metrics.get('Profit Before Tax', 0), 0.0)
    if taxation != 0 and profit_before_tax != 0:
        effective_tax_rate = (abs(taxation) / profit_before_tax) * 100
        sop_metrics['Effective Tax Rate'] = effective_tax_rate
        sop_source_terms['Effective Tax Rate'] = "Calculated from taxation and profit before tax"
        sop_calculations['Effective Tax Rate'] = f"({abs(taxation):,.2f} / {profit_before_tax:,.2f}) × 100 = {effective_tax_rate:.2f}%"
        print(f"  ✅ Effective Tax Rate: {effective_tax_rate:.2f}%")
    else:
        effective_tax_rate = 0.0
        sop_metrics['Effective Tax Rate'] = effective_tax_rate
    
    # Capital Expenditure (Quarterly) - already extracted as 'Capital Exp Qtrly'
    capital_expenditure = to_number(sop_metrics.get('Capital Exp Qtrly', 0), 0.0)
    if capital_expenditure == 0:
        # Try to calculate from components if not found directly
        capex_components = [
            'acquisition of property, plant & equipment',
            'acquisition of intangible assets',
            'purchase of property, plant and equipment',
            'additions to property, plant and equipment'
        ]
        capex_details = []
        for component in capex_components:
            value, source_term = find_value_in_dataframes_intelligent([component], dataframes)
            if value != 0:
                capital_expenditure += abs(value)  # Make positive (usually negative in cash flow)
                capex_details.append(f"{component}: {abs(value):,.2f}")
        
        if capital_expenditure > 0:
            sop_metrics['Capital Exp Qtrly'] = capital_expenditure
            sop_source_terms['Capital Exp Qtrly'] = "Calculated from components"
            sop_calculations['Capital Exp Qtrly'] = " + ".join(capex_details) if capex_details else "No components found"
    
    print(f"  ✅ Capital Exp Qtrly: {capital_expenditure:,.2f}")
    
    # Net Borrowings (Quarterly) - already extracted as 'Net Borrowings Qrtly'
    net_borrowings = to_number(sop_metrics.get('Net Borrowings Qrtly', 0), 0.0)
    if net_borrowings == 0:
        # Try to calculate from components if not found directly
        short_term_borrowings, _ = find_value_in_dataframes_intelligent(['short term borrowings', 'short-term borrowings'], dataframes)
        long_term_borrowings, _ = find_value_in_dataframes_intelligent(['long term borrowings', 'long-term borrowings'], dataframes)
        lease_payments, _ = find_value_in_dataframes_intelligent(['principal element of lease payment', 'lease payments'], dataframes)
        
        net_borrowings = short_term_borrowings + long_term_borrowings - lease_payments
        if net_borrowings != 0:
            sop_metrics['Net Borrowings Qrtly'] = net_borrowings
            sop_source_terms['Net Borrowings Qrtly'] = "Calculated from borrowings and lease payments"
            sop_calculations['Net Borrowings Qrtly'] = f"{short_term_borrowings:,.2f} + {long_term_borrowings:,.2f} - {lease_payments:,.2f} = {net_borrowings:,.2f}"
    
    print(f"  ✅ Net Borrowings Qrtly: {net_borrowings:,.2f}")
    
    print(f"\n📋 Knowledge-based SOP Extraction Summary:")
    print(f"  📊 Total metrics processed: {len(sop_metrics)}")
    found_count = sum(1 for v in sop_metrics.values() if v != 0)
    print(f"  ✅ Metrics found: {found_count}")
    print(f"  ❌ Metrics missing: {len(sop_metrics) - found_count}")
    
    return sop_metrics, sop_source_terms, sop_calculations

def extract_fallback_sops(dataframes, extracted_text=""):
    """Extract basic SOPs using knowledge when no company-specific config is available"""
    print("🔧 Extracting fallback SOPs using financial knowledge...")
    
    sop_metrics = {}
    sop_source_terms = {}
    sop_calculations = {}
    
    # Define basic SOP search terms based on common financial statement patterns
    basic_sops = {
        'Revenue': [
            'revenue', 'total revenue', 'income', 'total income', 'sales', 'turnover',
            'interest income', 'fee income', 'commission income', 'trading income'
        ],
        'Gross Profit': [
            'gross profit', 'gross income', 'net operating income', 'operating income'
        ],
        'Net Profit': [
            'net profit', 'profit for the period', 'profit for the year', 'net income',
            'profit/(loss) for the period', 'net earnings'
        ],
        'Operating Profit': [
            'operating profit', 'operating income', 'operating earnings', 'ebit',
            'operating profit before tax', 'operating profit before vat'
        ],
        'Profit Before Tax': [
            'profit before tax', 'earnings before tax', 'profit before taxation',
            'profit/ (loss) before income tax', 'pre-tax profit'
        ],
        'Taxation': [
            'taxation', 'tax expense', 'income tax', 'income tax expense', 'tax provision'
        ],
        'Total Assets': [
            'total assets', 'assets total'
        ],
        'Total Liabilities': [
            'total liabilities', 'liabilities total'
        ],
        'Cash': [
            'cash', 'cash and cash equivalents', 'cash equivalents'
        ],
        'Share Price': [
            'share price', 'price per share', 'market price', 'last traded',
            'market price of ordinary share', 'stock price'
        ],
        'Total Number of Issued Shares': [
            'total number of issued shares', 'number of shares', 'ordinary shares',
            'number of ordinary shares', 'issued shares'
        ],
        'Eps': [
            'earnings per share', 'eps', 'basic earnings per share', 'earning per share'
        ],
        'Book Value': [
            'book value', 'net asset value per share', 'net asset value per ordinary share'
        ]
    }
    
    # Use the enhanced find_value_in_dataframes from SOPEngine
    from sop_engine import SOPEngine
    temp_engine = SOPEngine()
    
    for sop_name, search_terms in basic_sops.items():
        print(f"🔍 Searching for {sop_name}...")
        value, source_term = temp_engine.find_value_in_dataframes(search_terms, dataframes)
        
        sop_metrics[sop_name] = value
        sop_source_terms[sop_name] = source_term if source_term else "Not found"
        sop_calculations[sop_name] = "Direct extraction from financial statements"
        
        if value != 0:
            print(f"  ✅ Found: {value:,.2f} (from: {source_term})")
        else:
            print(f"  ❌ Not found")
    
    # Calculate some basic derived metrics
    print("\n🧮 Calculating derived metrics...")
    
    # Market Capitalization = Share Price × Total Number of Issued Shares
    share_price = sop_metrics.get('Share Price', 0)
    total_shares = sop_metrics.get('Total Number of Issued Shares', 0)
    if share_price > 0 and total_shares > 0:
        market_cap = share_price * total_shares
        sop_metrics['Market Capitalization'] = market_cap
        sop_source_terms['Market Capitalization'] = "Calculated from share price and shares"
        sop_calculations['Market Capitalization'] = f"{share_price:,.2f} × {total_shares:,.0f} = {market_cap:,.2f}"
        print(f"  ✅ Market Capitalization: {market_cap:,.2f}")
    
    # Tax Rate = (Taxation / Revenue) × 100
    taxation = sop_metrics.get('Taxation', 0)
    revenue = sop_metrics.get('Revenue', 0)
    if taxation != 0 and revenue != 0:
        tax_rate = (abs(taxation) / revenue) * 100
        sop_metrics['Tax Rate'] = tax_rate
        sop_source_terms['Tax Rate'] = "Calculated from taxation and revenue"
        sop_calculations['Tax Rate'] = f"({abs(taxation):,.2f} / {revenue:,.2f}) × 100 = {tax_rate:.2f}%"
        print(f"  ✅ Tax Rate: {tax_rate:.2f}%")
    
    # Effective Tax Rate = (Taxation / Profit Before Tax) × 100
    profit_before_tax = sop_metrics.get('Profit Before Tax', 0)
    if taxation != 0 and profit_before_tax != 0:
        eff_tax_rate = (abs(taxation) / profit_before_tax) * 100
        sop_metrics['Effective Tax Rate'] = eff_tax_rate
        sop_source_terms['Effective Tax Rate'] = "Calculated from taxation and profit before tax"
        sop_calculations['Effective Tax Rate'] = f"({abs(taxation):,.2f} / {profit_before_tax:,.2f}) × 100 = {eff_tax_rate:.2f}%"
        print(f"  ✅ Effective Tax Rate: {eff_tax_rate:.2f}%")
    
    print(f"\n📊 Fallback SOP extraction completed: {len(sop_metrics)} metrics")
    return sop_metrics, sop_source_terms, sop_calculations

def add_sop_metrics_to_excel(excel_file_path, company_type=None, extracted_text="", pdf_path=None):
    """Add SOP metrics sheet to existing Excel file using AI knowledge-based extraction"""
    print(f"📊 Reading financial statements from: {excel_file_path}")
    
    # Read all sheets from the Excel file
    xl = pd.ExcelFile(excel_file_path)
    dataframes = []
    
    # Convert Excel sheets back to dataframes for processing
    for sheet_name in xl.sheet_names:
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        if not df.empty:
            dataframes.append(df)
            print(f"  📋 Loaded {sheet_name}: {len(df)} rows")
    
    if not dataframes:
        raise Exception("No data found in Excel file")
    
    # Use knowledge-based SOP extraction (no configuration files needed)
    print(f"🧠 Using AI knowledge-based SOP extraction...")
    from main import extract_sop_metrics_from_dataframes_knowledge_based
    sop_metrics, sop_source_terms, sop_calculations = extract_sop_metrics_from_dataframes_knowledge_based(
        dataframes,
        extracted_text
    )
    
    print(f"📊 Calculated {len(sop_metrics)} SOP metrics using AI knowledge")
    
    # Add SOP metrics sheet to the existing Excel file
    print("💾 Adding SOP metrics sheet to Excel file...")
    
    # Read existing file to preserve other sheets
    with pd.ExcelWriter(excel_file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        # Create SOP dataframe as key/value only
        sop_data = []
        for metric in sop_metrics.keys():
            sop_data.append({
                "Metric": metric,
                "Value": fmt_value(sop_metrics[metric])
            })
        sop_df = pd.DataFrame(sop_data, columns=["Metric", "Value"]) 
        sop_df.to_excel(writer, sheet_name="SOP_Metrics", index=False)
        
        # Apply formatting
        worksheet = writer.sheets["SOP_Metrics"]
        apply_excel_formatting(worksheet, sop_df)
    
    print(f"✅ SOP metrics sheet added with {len(sop_metrics)} metrics")

def detect_company_config_from_pdf(pdf_path):
    """Detect company-specific SOP configuration based on PDF filename"""
    
    # Extract base filename without extension
    base_filename = os.path.splitext(os.path.basename(pdf_path))[0]
    
    # Check if a company-specific JSON config exists
    company_config_path = os.path.join("sop_configs", f"{base_filename}.json")
    
    if os.path.exists(company_config_path):
        print(f"🎯 Found company-specific configuration: {company_config_path}")
        return base_filename, company_config_path
    else:
        print(f"📋 No company-specific config found for {base_filename}")
        print(f"🔍 Looking for: {company_config_path}")
        return None, None

def detect_or_select_company_type(pdf_path=None):
    """Detect or ask user to select company type for SOP configuration"""
    
    # First try to find company-specific config if PDF path is provided
    if pdf_path:
        company_id, config_path = detect_company_config_from_pdf(pdf_path)
        if company_id and config_path:
            return company_id
        else:
            # No company-specific config found, use fallback SOP extraction
            print("📋 No company-specific config found, will use fallback SOP extraction")
            return "fallback"
    
    # Fallback path deprecated; return fallback directly
    available_configs = []
    
    if not available_configs:
        print("⚠️ No SOP configurations found. Using fallback SOP extraction.")
        return "fallback"
    
    print("\n🏢 Available Company Types:")
    for i, config_type in enumerate(available_configs, 1):
        print(f"  {i}. {config_type.title()}")
    
    # For now, default to bank (later we can add interactive selection)
    print(f"\n🔧 Auto-selecting 'bank' configuration (default)")
    print("💡 To use a different configuration, modify the code or create custom configs")
    
    return "bank"

# =====================
# Step 5: Main Processing Function
# =====================
def process_annual_report(pdf_path, output_filename=None):
    """
    Process an annual report PDF and extract financial statements to Excel
    
    Args:
        pdf_path (str): Path to the PDF file
        output_filename (str): Optional custom output filename
    
    Returns:
        str: Path to the generated Excel file
    """
    if not os.path.exists(pdf_path):
        print(f"❌ Error: PDF file not found at {pdf_path}")
        return None
    
    print(f"🔄 Processing PDF: {pdf_path}")
    text = extract_pdf_text(pdf_path)
    
    if not text:
        print("❌ PDF text extraction failed.")
        return None
    
    print("✅ PDF text extracted successfully")
    print("🔄 Extracting financial statements...")
    
    tables = {
        "Profit or Loss": extract_profit_or_loss(text),
        "Comprehensive Income": extract_comprehensive_income(text),
        "Financial Position": extract_financial_position(text),
        "Changes in Equity": extract_changes_in_equity(text),
        "Cash Flows": extract_cash_flows(text)
    }
    
    # Generate output filename if not provided
    if not output_filename:
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_filename = f"{base_name}_Statements.xlsx"
    
    # STEP 1: Save the 5 financial statements first
    print("💾 Step 1: Saving financial statements to Excel...")
    excel_file = save_all_to_excel(tables, output_filename)
    print(f"✅ Financial statements saved: {excel_file}")
    
    # STEP 2: Add SOP via LLM (combined flow)
    try:
        print("🧠 Step 2: Deriving SOP via LLM and appending to Excel...")
        # Default model can be overridden via env var SOP_OPENROUTER_MODEL
        sop_model = os.getenv("SOP_OPENROUTER_MODEL", "google/gemini-2.5-flash")
        add_sop_metrics_via_llm(excel_file, model=sop_model, temperature=0)
        print("✅ SOP_Metrics sheet added")
    except Exception as e:
        print(f"⚠️ SOP addition failed (continuing with statements only): {e}")
    
    print(f"🎉 Processing complete! Excel file created: {excel_file}")
    return excel_file

# =====================
# Step 6: Run Everything
# =====================
if __name__ == "__main__":
    # Update this path to your PDF file
    pdf_path = "bairaha.pdf"  # You can also use full path like r"D:\path\to\your\file.pdf"
    
    # Process the annual report
    result_file = process_annual_report(pdf_path)
    
    if result_file:
        print(f"\n✅ Success! Check your Excel file: {result_file}")
        print("📊 SOP Metrics have been added to the 'SOP_Metrics' sheet in the Excel file")
    else:
        print("\n❌ Processing failed. Please check the PDF path and try again.")
